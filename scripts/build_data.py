#!/usr/bin/env python3
"""
Lee el Excel de peso y genera data/peso.json para el dashboard de salud.

Reglas de lectura de la hoja de peso:
- Fila 1, col A: nombre del mes (p. ej. "Agosto").
- Fila 2: cabeceras (Día, Peso, Eat).
- Fila 3 en adelante: datos. Col A = día del mes, col B = peso (puede estar vacío),
  col C = marca "Eat" (de momento se ignora).
- Hay días sin peso (huecos) en medio: NO parar en el primer hueco; recorrer todo.
- La última fila con un número suelto (sin día) es la media del mes: ignorarla.
"""

import json
import sys
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

# --- Configuración -------------------------------------------------------
# Ruta del Excel de peso en el Mac (cámbiala solo si mueves el archivo).
EXCEL_PATH = (
    "/Users/davidrochgarcia/Library/CloudStorage/"
    "GoogleDrive-davidrochgarcia@gmail.com/My Drive/1. dOS/Salud/Peso.xlsx"
)
# Dónde se escribe el JSON (relativo a la raíz del proyecto).
OUTPUT_PATH = "data/peso.json"

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}


def _sin_acentos(texto: str) -> str:
    t = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in t if not unicodedata.combining(c)).strip().lower()


def leer_peso(excel_path: str) -> dict:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.worksheets[0]

    # Mes (fila 1, col A) -> número de mes.
    titulo = ws.cell(row=1, column=1).value
    mes_txt = _sin_acentos(str(titulo)) if titulo else ""
    mes_num = MESES.get(mes_txt.split()[0]) if mes_txt else None
    anio = datetime.now().year

    registros = []
    for fila in ws.iter_rows(min_row=3, values_only=True):
        dia = fila[0] if len(fila) > 0 else None
        peso = fila[1] if len(fila) > 1 else None
        # Solo filas con día numérico entre 1 y 31 y con peso numérico.
        if not isinstance(dia, (int, float)):
            continue
        dia = int(dia)
        if dia < 1 or dia > 31:
            continue
        if not isinstance(peso, (int, float)):
            continue
        if mes_num:
            fecha = f"{anio:04d}-{mes_num:02d}-{dia:02d}"
        else:
            fecha = None
        registros.append({"date": fecha, "day": dia, "weight": round(float(peso), 1)})

    registros.sort(key=lambda r: r["day"])
    return {
        "mes_txt": str(titulo).strip() if titulo else "",
        "mes_num": mes_num,
        "anio": anio,
        "registros": registros,
    }


def dias_entre(fecha_a: str, fecha_b: str) -> int:
    a = datetime.strptime(fecha_a, "%Y-%m-%d")
    b = datetime.strptime(fecha_b, "%Y-%m-%d")
    return abs((a - b).days)


def construir_json(datos: dict) -> dict:
    regs = datos["registros"]
    if not regs:
        return {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "unit": "kg",
            "month": f'{datos["mes_txt"]} {datos["anio"]}'.strip(),
            "current": None,
            "series": [],
            "note": "Sin registros de peso este mes.",
        }

    pesos = [r["weight"] for r in regs]
    current = regs[-1]
    previous = regs[-2] if len(regs) >= 2 else None

    delta_prev = round(current["weight"] - previous["weight"], 1) if previous else None

    # Variación vs el MISMO día de la semana pasada = exactamente 7 días antes.
    # null si no hay registro ese día.
    delta_week = None
    ref_week = None
    if current["date"]:
        for r in regs[:-1]:
            if r["date"] and dias_entre(current["date"], r["date"]) == 7:
                ref_week = r
                delta_week = round(current["weight"] - r["weight"], 1)
                break

    # Medias por SEMANA NATURAL (lunes-domingo):
    #   - "this": semana en curso, del lunes a hoy (semana a fecha).
    #   - "last": semana natural anterior completa (lunes a domingo).
    # Robusto ante huecos: promedia los días que haya en cada rango.
    def media_rango(desde, hasta):
        vals = []
        for r in regs:
            if not r["date"]:
                continue
            dt = datetime.strptime(r["date"], "%Y-%m-%d")
            if desde <= dt <= hasta:
                vals.append(r["weight"])
        if not vals:
            return {"avg": None, "count": 0, "_sum": 0.0}
        return {"avg": round(sum(vals) / len(vals), 2), "count": len(vals),
                "_sum": sum(vals)}

    week = None
    if current["date"]:
        cur_dt = datetime.strptime(current["date"], "%Y-%m-%d")
        this_start = cur_dt - timedelta(days=cur_dt.weekday())   # lunes de esta semana
        last_start = this_start - timedelta(days=7)
        last_end = this_start - timedelta(days=1)                # domingo pasado

        this_w = media_rango(this_start, cur_dt)
        last_w = media_rango(last_start, last_end)
        delta_w = None
        if this_w["avg"] is not None and last_w["avg"] is not None:
            raw_this = this_w["_sum"] / this_w["count"]
            raw_last = last_w["_sum"] / last_w["count"]
            delta_w = round(raw_this - raw_last, 2)
        for w in (this_w, last_w):
            w.pop("_sum", None)
        week = {
            "this": this_w, "last": last_w, "delta": delta_w,
            "this_start": this_start.strftime("%Y-%m-%d"),
            "last_start": last_start.strftime("%Y-%m-%d"),
            "last_end": last_end.strftime("%Y-%m-%d"),
        }

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "unit": "kg",
        "month": f'{datos["mes_txt"]} {datos["anio"]}'.strip(),
        "current": current,
        "previous": previous,
        "delta_vs_previous": delta_prev,
        "delta_vs_week": delta_week,
        "reference_week": ref_week,
        "week": week,
        "stats": {
            "count": len(regs),
            "min": min(pesos),
            "max": max(pesos),
            "avg": round(sum(pesos) / len(pesos), 1),
        },
        "series": regs,
    }


def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else EXCEL_PATH
    out_path = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_PATH

    datos = leer_peso(excel_path)
    resultado = construir_json(datos)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    # Resumen legible.
    c = resultado.get("current")
    if c:
        print(f"Mes: {resultado['month']}")
        print(f"Registros: {resultado['stats']['count']}")
        print(f"Peso actual: {c['weight']} kg (día {c['day']}, {c['date']})")
        print(f"Variación vs día anterior: {resultado['delta_vs_previous']}")
        print(f"Variación vs ~semana: {resultado['delta_vs_week']}")
        print(f"Media/min/max: {resultado['stats']['avg']} / "
              f"{resultado['stats']['min']} / {resultado['stats']['max']}")
    else:
        print("Sin registros.")
    print(f"\nEscrito: {out_path}")


if __name__ == "__main__":
    main()
