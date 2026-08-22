#!/usr/bin/env python3
"""
Lee el Excel de entreno y genera data/deporte.json para el dashboard.

Métrica: "días activos" de la semana natural en curso (lunes→hoy).
- Cada "Día 1…7" del Excel es un HUECO/tipo de deporte (no un día fijo).
- Un hueco cuenta como "día activo" si en la columna de esta semana hay algún
  valor escrito; un guion "-" o vacío = no.
- Días activos = nº de huecos con algo apuntado esta semana.
- Denominador = días naturales transcurridos (lunes→hoy).
"""

import json
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

EXCEL_PATH = (
    "/Users/davidrochgarcia/Library/CloudStorage/"
    "GoogleDrive-davidrochgarcia@gmail.com/My Drive/1. dOS/Salud/Entrenamiento/Entreno 2026.xlsx"
)
OUTPUT_PATH = "data/deporte.json"

MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
         "agosto", "septiembre", "octubre", "noviembre", "diciembre"]


def is_value(v):
    if v is None:
        return False
    s = str(v).strip()
    return s not in ("", "-", "–", "—")


def elegir_hoja(wb, hoy):
    """Hoja del mes en curso ('Entreno <Mes> <año>'); si no, la que tenga la
    columna de esta semana; si no, la primera."""
    objetivo = f"Entreno {MESES[hoy.month - 1].capitalize()} {hoy.year}"
    if objetivo in wb.sheetnames:
        return wb[objetivo]
    lunes = hoy - timedelta(days=hoy.weekday())
    for ws in wb.worksheets:
        for c in ws[2]:
            if isinstance(c.value, datetime) and c.value.date() == lunes.date():
                return ws
    return wb.worksheets[0]


def col_semana(ws, hoy):
    """Columna cuya cabecera (fila 2) es el lunes de esta semana; si no existe,
    la fecha más reciente que sea <= hoy."""
    lunes = (hoy - timedelta(days=hoy.weekday())).date()
    fechas = []
    for c in ws[2]:
        if isinstance(c.value, datetime):
            fechas.append((c.value.date(), c.column))
    for d, col in fechas:
        if d == lunes:
            return col, d
    anteriores = [(d, col) for d, col in fechas if d <= hoy.date()]
    if anteriores:
        d, col = max(anteriores, key=lambda x: x[0])
        return col, d
    return None, None


def contar(ws, col):
    """Nº de bloques 'Día' con algún valor en la columna dada."""
    activos = 0
    total_huecos = 0
    cur_tiene = None
    for r in ws.iter_rows(min_row=3, values_only=False):
        a = r[0].value
        a_str = str(a).strip() if a is not None else ""
        if a_str.upper().startswith("NOTAS"):
            break
        if a_str.lower().startswith("día") or a_str.lower().startswith("dia"):
            # nuevo bloque "Día N"
            if cur_tiene:
                activos += 1
            cur_tiene = False
            total_huecos += 1
        elif cur_tiene is not None:
            # fila de un bloque (la col A puede estar vacía: deporte libre)
            cell = r[col - 1].value if len(r) >= col else None
            if is_value(cell):
                cur_tiene = True
    if cur_tiene:
        activos += 1
    return activos, total_huecos


def main():
    excel = sys.argv[1] if len(sys.argv) > 1 else EXCEL_PATH
    out = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_PATH
    hoy = datetime.now()

    wb = openpyxl.load_workbook(excel, data_only=True)
    ws = elegir_hoja(wb, hoy)
    col, fecha_col = col_semana(ws, hoy)

    transcurridos = hoy.weekday() + 1  # lunes=1 … domingo=7
    if col is None:
        activos_raw = 0
    else:
        activos_raw, _ = contar(ws, col)

    activos = min(activos_raw, transcurridos)  # nunca "X de menos que X"
    total = 7

    patron = (["active"] * activos
              + ["rest"] * (transcurridos - activos)
              + ["future"] * (total - transcurridos))

    lunes = (hoy - timedelta(days=hoy.weekday())).date()
    res = {
        "generated_at": hoy.isoformat(timespec="seconds"),
        "week_start": lunes.isoformat(),
        "dias_activos": activos,
        "dias_transcurridos": transcurridos,
        "total_semana": total,
        "slots_marcados": activos_raw,
        "pattern": patron,
        "sheet": ws.title,
        "week_col_date": fecha_col.isoformat() if fecha_col else None,
    }

    Path(out).parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(res, f, ensure_ascii=False, indent=2)

    print(f"Hoja: {ws.title} · semana del {lunes} (col {fecha_col})")
    print(f"Días activos: {activos} de {transcurridos}  (huecos marcados: {activos_raw})")
    print(f"Patrón: {patron}")
    print(f"Escrito: {out}")


if __name__ == "__main__":
    main()
